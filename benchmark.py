"""
Cis-GS Benchmarking Suite  v1.2.0
-----------------------------------
Pipeline:
  1. Promoter Extraction   — 4 genomes × 2000 bp
  2. Motif Search          — 2 manuscript motifs on each promoter set
  3. Co-expression         — 3 species:
                             Arachis gnm1  (67k genes × 54 samples, raw counts)
                             Oryza sativa  (27.8k genes × 80 samples, TPM)
                             Medicago trunc.(27.1k genes × 27 samples, TPM)
  4. k-means Clustering    — elbow method + k-means at optimal k (all 3 species)
  5. KEGG Enrichment       — 3 organisms (ahf, osa, mtr):
                             network fetch, disk cache, ID Convert (ahf),
                             module enrichment, scaling

Run:   python benchmark.py
Output: benchmark_results.csv  +  benchmark_results.xlsx
"""

import sys, time, platform
from pathlib import Path
from datetime import datetime

import numpy as np
import pandas as pd

sys.path.insert(0, str(Path(__file__).parent))

# ── Hardware fingerprint ────────────────────────────────────────────────────
import psutil
CPU_NAME    = platform.processor()
CPU_CORES   = psutil.cpu_count(logical=False)
CPU_THREADS = psutil.cpu_count(logical=True)
RAM_GB      = round(psutil.virtual_memory().total / 1e9, 1)
OS_NAME     = f"{platform.system()} {platform.release()}"
PY_VER      = platform.python_version()

print("=" * 70)
print("  Cis-GS Benchmark Suite  v1.2.0")
print(f"  {datetime.now().strftime('%Y-%m-%d  %H:%M:%S')}")
print(f"  OS     : {OS_NAME}")
print(f"  CPU    : {CPU_NAME}")
print(f"  Cores  : {CPU_CORES} physical / {CPU_THREADS} logical")
print(f"  RAM    : {RAM_GB} GB")
print(f"  Python : {PY_VER}")
print("=" * 70)

# ── Timing harness ──────────────────────────────────────────────────────────
_proc = psutil.Process()

def run_timed(label, fn, *args, n_reps=3, **kwargs):
    """
    All reps clean (no tracemalloc overhead).
    Peak RAM estimated via psutil RSS delta.
    Returns (mean_s, sd_s, peak_ram_mb, last_result)
    """
    times, peak_mb = [], 0.0
    result = None
    for i in range(n_reps):
        ram_before = _proc.memory_info().rss / 1e6
        t0 = time.perf_counter()
        result = fn(*args, **kwargs)
        elapsed = time.perf_counter() - t0
        ram_after = _proc.memory_info().rss / 1e6
        peak_mb = max(peak_mb, ram_after - ram_before)
        times.append(elapsed)
        print(f"    rep {i+1}/{n_reps}:  {elapsed:.2f} s")
    mean_t = float(np.mean(times))
    sd_t   = float(np.std(times, ddof=1)) if n_reps > 1 else 0.0
    print(f"  + {label}:  {mean_t:.2f} +/- {sd_t:.3f} s  |  peak RAM delta {peak_mb:.0f} MB")
    return mean_t, sd_t, peak_mb, result

results = []

def record(function, dataset, parameter, genome_mb,
           n_input, n_output, mean_t, sd_t, peak_mb,
           throughput="", unit=""):
    results.append({
        "function":        function,
        "dataset":         dataset,
        "parameter":       parameter,
        "genome_size_mb":  genome_mb,
        "n_input":         n_input,
        "n_output":        n_output,
        "mean_time_s":     round(mean_t, 3),
        "sd_time_s":       round(sd_t,   4),
        "peak_ram_mb":     round(peak_mb, 1),
        "throughput":      throughput,
        "throughput_unit": unit,
    })


# ══════════════════════════════════════════════════════════════════════════
# CONFIG
# ══════════════════════════════════════════════════════════════════════════

PROMOTER_LEN = 2000
OUT_DIR = Path("benchmark_out"); OUT_DIR.mkdir(exist_ok=True)

# ── 4 genomes for promoter extraction + motif search ──────────────────────
GENOMES = [
    {
        "label":   "Oryza sativa IRGSP-1.0",
        "size_mb": 362,
        "genome":  Path(r"C:\Users\ayush\OneDrive\Documents\rice_genome_data\IRGSP-1.0_genome.fasta"),
        "gff":     Path(r"C:\Users\ayush\OneDrive\Documents\rice_genome_data\IRGSP-1.0_annotation\locus.gff"),
        "key":     "rice",
    },
    {
        "label":   "Medicago truncatula Mt5.0",
        "size_mb": 421,
        "genome":  Path(r"C:\Users\ayush\OneDrive\Documents\Cis-GS-V3\workspace\genomes\Medicago_genome.fasta"),
        "gff":     Path(r"C:\Users\ayush\OneDrive\Documents\Cis-GS-V3\workspace\annotations\Medicago_annotation.gff3"),
        "key":     "medicago",
    },
    {
        "label":   "Arachis hypogaea gnm1 (Phytozome)",
        "size_mb": 2500,
        "genome":  Path(r"C:\Users\ayush\OneDrive\Documents\Cis-GS-V3\workspace\genomes\Arachis_gnm1_genome.fasta"),
        "gff":     Path(r"C:\Users\ayush\OneDrive\Documents\Cis-GS-V3\workspace\annotations\Arachis_gnm1_annotations.gff3"),
        "key":     "arachis_gnm1",
    },
    {
        "label":   "Arachis hypogaea gnm2 (NCBI RefSeq)",
        "size_mb": 2500,
        "genome":  Path(r"C:\Users\ayush\OneDrive\Documents\Cis-GS-V3\workspace\genomes\Arachis_gnm2_genome.fasta"),
        "gff":     Path(r"C:\Users\ayush\OneDrive\Documents\Cis-GS-V3\workspace\annotations\Arachis_gnm2_annotation.gff3"),
        "key":     "arachis_gnm2",
    },
]

# ── Manuscript TF motifs ───────────────────────────────────────────────────
TF_MOTIFS = [
    ("CYC-RE_RAM1", "TGGGCCGGCCCA"),
    ("CYC-RE_NIN",  "NGCCATGTGGCN"),
]

# ── Expression data paths ──────────────────────────────────────────────────
TISSUE_CSV    = Path(r"C:\Users\ayush\Downloads\Arahy_all_tissue_htseq_count.csv")
RICE_TPM_CSV  = Path(r"C:\Users\ayush\Downloads\expression_data\rice_tissue_TPM_RAPids.csv")
MEDTR_TPM_CSV = Path(r"C:\Users\ayush\Downloads\expression_data\medicago_tissue_TPM_MedtrIDs.csv")

# ── KEGG: Arachis real hit genes (for ID Convert benchmark) ───────────────
HITS_CSV      = Path(r"C:\Users\ayush\Downloads\arachis hits.csv")

# ── Arahy -> LOC crosswalk (for Arachis KEGG module enrichment) ───────────
CROSSWALK_XLSX = Path(r"C:\Users\ayush\OneDrive\Desktop\arahy_LOC_crosswalk.xlsx")


# ── Co-expression species config ───────────────────────────────────────────
#    corr_subset : number of most-variable genes used for the correlation matrix
COEXPR_SPECIES = [
    {
        "label":       "Arachis hypogaea gnm1",
        "key":         "arachis",
        "kegg_org":    "ahf",
        "csv":         TISSUE_CSV,
        "id_starts":   ("Arahy.", "arahy."),
        "corr_subset": 5000,
        "description": "raw HTSeq counts, log2(n+1)",
    },
    {
        "label":       "Oryza sativa IRGSP-1.0",
        "key":         "rice",
        "kegg_org":    "osa",
        "csv":         RICE_TPM_CSV,
        "id_starts":   ("Os",),
        "corr_subset": 5000,
        "description": "TPM (80 samples, 34 tissues), log2(TPM+1)",
    },
    {
        "label":       "Medicago truncatula Mt5.0",
        "key":         "medicago",
        "kegg_org":    "mtr",
        "csv":         MEDTR_TPM_CSV,
        "id_starts":   ("Medtr",),
        "corr_subset": 3000,
        "description": "TPM (27 samples, 9 tissues), log2(TPM+1)",
    },
]


# ══════════════════════════════════════════════════════════════════════════
# 1. PROMOTER EXTRACTION  (4 genomes)
# ══════════════════════════════════════════════════════════════════════════
print("\n" + "-" * 70)
print("  STEP 1 -- Promoter Extraction  (2000 bp upstream TSS)")
print("-" * 70)

from cis_gs.app import extract_promoters

promoter_fastas = {}   # key -> Path

for g in GENOMES:
    if not g["genome"].exists() or not g["gff"].exists():
        print(f"\n  SKIP (file not found): {g['label']}")
        continue

    out_fa  = OUT_DIR / f"{g['key']}_promoters.fasta"
    out_csv = OUT_DIR / f"{g['key']}_promoter_map.csv"
    print(f"\n  {g['label']}  ({g['size_mb']} MB genome)")

    mean_t, sd_t, peak_mb, (n_genes, n_written) = run_timed(
        g["label"], extract_promoters,
        g["genome"], g["gff"], out_fa, out_csv, PROMOTER_LEN,
        n_reps=3,
    )
    promoter_fastas[g["key"]] = out_fa
    record(
        function  = "Promoter Extraction",
        dataset   = g["label"],
        parameter = f"{PROMOTER_LEN} bp",
        genome_mb = g["size_mb"],
        n_input   = n_genes, n_output=n_written,
        mean_t=mean_t, sd_t=sd_t, peak_mb=peak_mb,
        throughput=round(n_written / mean_t, 1) if mean_t > 0 else "",
        unit="genes/s",
    )


# ══════════════════════════════════════════════════════════════════════════
# 2. MOTIF SEARCH  (4 genomes, 2 manuscript motifs)
# ══════════════════════════════════════════════════════════════════════════
print("\n" + "-" * 70)
print("  STEP 2 -- Motif Search  (CYC-RE_RAM1 + CYC-RE_NIN)")
print("-" * 70)

from cis_gs.app import scan_fasta_for_motifs

for g in GENOMES:
    key = g["key"]
    if key not in promoter_fastas:
        print(f"\n  SKIP (no promoter FASTA): {g['label']}")
        continue
    prom_fa = promoter_fastas[key]
    if not prom_fa.exists():
        print(f"\n  SKIP (promoter FASTA missing): {g['label']}")
        continue

    n_promoters = sum(1 for line in open(prom_fa) if line.startswith(">"))
    print(f"\n  {g['label']}  ({n_promoters:,} promoters, {len(TF_MOTIFS)} motifs)")

    mean_t, sd_t, peak_mb, hit_df = run_timed(
        g["label"], scan_fasta_for_motifs, prom_fa, TF_MOTIFS, n_reps=3,
    )
    n_hits = len(hit_df) if hit_df is not None and not hit_df.empty else 0
    record(
        function  = "Motif Search",
        dataset   = g["label"],
        parameter = "CYC-RE_RAM1 + CYC-RE_NIN",
        genome_mb = g["size_mb"],
        n_input   = n_promoters, n_output=n_hits,
        mean_t=mean_t, sd_t=sd_t, peak_mb=peak_mb,
        throughput=round(n_promoters / mean_t, 1) if mean_t > 0 else "",
        unit="promoters/s",
    )


# ══════════════════════════════════════════════════════════════════════════
# 3. CO-EXPRESSION CLUSTERING  (3 species)
#    3a. Normalisation
#    3b. Pearson correlation matrix (top N variable genes)
#    3c. Louvain module detection
# ══════════════════════════════════════════════════════════════════════════
print("\n" + "-" * 70)
print("  STEP 3 -- Co-expression Clustering  (3 species)")
print("-" * 70)

from cis_gs.app import normalize_expression_data, calculate_correlation_matrix, detect_coexpression_modules

# Per-species results (used by STEPS 4 and 5)
coexpr_sub         = {}   # key -> normalised subset DataFrame (top N variable)
coexpr_communities = {}   # key -> Louvain communities dict
coexpr_corr        = {}   # key -> correlation matrix
n_samples_map      = {}   # key -> number of samples

for sp in COEXPR_SPECIES:
    key    = sp["key"]
    label  = sp["label"]
    csv    = sp["csv"]
    starts = sp["id_starts"]
    N      = sp["corr_subset"]

    if not csv.exists():
        print(f"\n  SKIP -- CSV not found: {csv}")
        continue

    print(f"\n  {label}  ({sp['description']})")
    raw = pd.read_csv(csv, index_col=0)

    # Drop metadata columns if present (e.g. 'biotype', 'source')
    drop_cols = [c for c in raw.columns if c.lower() in ("biotype", "source",
                                                           "functional_annotation")]
    if drop_cols:
        raw = raw.drop(columns=drop_cols)

    # Keep rows with valid gene IDs
    raw = raw[raw.index.astype(str).str.startswith(starts)]
    raw = raw.loc[(raw > 0).any(axis=1)]
    n_genes, n_samp = raw.shape
    n_samples_map[key] = n_samp
    print(f"  Matrix: {n_genes:,} genes x {n_samp} samples")

    # ── 3a. Normalisation ──────────────────────────────────────────────────
    print(f"  3a. log2 normalisation")
    mean_t, sd_t, peak_mb, norm_df = run_timed(
        f"Normalisation [{label}]", normalize_expression_data, raw, "log2", n_reps=5,
    )
    record(
        function  = "Expression Normalisation",
        dataset   = label,
        parameter = "log2(n+1)",
        genome_mb = "", n_input=n_genes, n_output=n_genes,
        mean_t=mean_t, sd_t=sd_t, peak_mb=peak_mb,
        throughput=round(n_genes / mean_t), unit="genes/s",
    )

    # ── 3b. Correlation matrix (top N variable genes) ─────────────────────
    print(f"  3b. Pearson correlation matrix  (top {N:,} variable genes)")
    var_genes = norm_df.var(axis=1).nlargest(N).index
    sub = norm_df.loc[var_genes]
    coexpr_sub[key] = sub

    mean_t, sd_t, peak_mb, corr = run_timed(
        f"Correlation [{label}]", calculate_correlation_matrix, sub, "pearson", n_reps=3,
    )
    coexpr_corr[key] = corr
    record(
        function  = "Correlation Matrix",
        dataset   = f"{label} (top {N:,} variable genes)",
        parameter = "Pearson",
        genome_mb = "", n_input=N, n_output=f"{N}x{N}",
        mean_t=mean_t, sd_t=sd_t, peak_mb=peak_mb,
    )

    # ── 3c. Louvain module detection ──────────────────────────────────────
    print(f"  3c. Louvain module detection  (r >= 0.70)")
    mean_t, sd_t, peak_mb, (comms, G) = run_timed(
        f"Louvain [{label}]", detect_coexpression_modules, corr, 0.70, "louvain", n_reps=3,
    )
    coexpr_communities[key] = comms
    n_mod = len(set(comms.values()))
    print(f"  -> {n_mod} co-expression modules detected")
    record(
        function  = "Co-expression Clustering (Louvain)",
        dataset   = f"{label} (top {N:,} variable genes)",
        parameter = "r >= 0.70",
        genome_mb = "", n_input=N, n_output=n_mod,
        mean_t=mean_t, sd_t=sd_t, peak_mb=peak_mb,
        throughput=round(N / mean_t), unit="genes/s",
    )


# ══════════════════════════════════════════════════════════════════════════
# 4. K-MEANS CLUSTERING  (3 species)
#    4a. Elbow method  (k = 2 ... 15)
#    4b. k-means at optimal k
# ══════════════════════════════════════════════════════════════════════════
print("\n" + "-" * 70)
print("  STEP 4 -- k-means Clustering  (3 species)")
print("-" * 70)

from cis_gs.app import elbow_method_kmeans, kmeans_temporal_clustering

kmeans_communities = {}   # key -> communities dict

for sp in COEXPR_SPECIES:
    key   = sp["key"]
    label = sp["label"]
    N     = sp["corr_subset"]
    sub   = coexpr_sub.get(key)

    if sub is None:
        print(f"\n  SKIP -- expression subset unavailable: {label}")
        continue

    print(f"\n  {label}  ({len(sub):,} genes x {n_samples_map.get(key, '?')} samples)")

    # ── 4a. Elbow method ──────────────────────────────────────────────────
    print(f"  4a. Elbow method  (k = 2 ... 15)")
    mean_t, sd_t, peak_mb, elbow_res = run_timed(
        f"Elbow [{label}]", elbow_method_kmeans, sub, 15, n_reps=3,
    )
    optimal_k = elbow_res["optimal_k"]
    print(f"  -> Optimal k: {optimal_k}")
    record(
        function  = "k-means Elbow Method",
        dataset   = f"{label} (top {N:,} variable genes)",
        parameter = "k = 2-15, second-derivative elbow",
        genome_mb = "", n_input=len(sub), n_output=f"optimal k={optimal_k}",
        mean_t=mean_t, sd_t=sd_t, peak_mb=peak_mb,
    )

    # ── 4b. k-means at optimal k ──────────────────────────────────────────
    print(f"  4b. k-means  (k={optimal_k}, n_init=50)")
    mean_t, sd_t, peak_mb, km_res = run_timed(
        f"k-means k={optimal_k} [{label}]", kmeans_temporal_clustering,
        sub, optimal_k, 50, n_reps=3,
    )
    kmeans_communities[key] = km_res["cluster_labels"].to_dict()
    record(
        function  = "k-means Clustering",
        dataset   = f"{label} (top {N:,} variable genes)",
        parameter = f"k={optimal_k}, n_init=50, z-score normalised",
        genome_mb = "", n_input=len(sub), n_output=optimal_k,
        mean_t=mean_t, sd_t=sd_t, peak_mb=peak_mb,
        throughput=round(len(sub) / mean_t), unit="genes/s",
    )


# ══════════════════════════════════════════════════════════════════════════
# 5. KEGG ENRICHMENT  (organisms: ahf, osa, mtr)
#    5a. First-time pathway fetch   (network)
#    5b. Cached pathway fetch       (disk)
#    5c. ID Convert                 (ahf — Arachis LOC IDs)
#    5d. Enrichment per Louvain module
#    5e. Enrichment per k-means cluster
#    5f. Scaling test
# ══════════════════════════════════════════════════════════════════════════
print("\n" + "-" * 70)
print("  STEP 5 -- KEGG Enrichment  (ahf / osa / mtr)")
print("-" * 70)

from cis_gs.enrichment.kegg import KEGGClient, KEGGEnricher
from cis_gs.enrichment.core import hypergeometric_enrichment

_kegg_client = KEGGClient()
_kegg_cache  = _kegg_client.cache_dir

KEGG_ORGS = [
    ("ahf", "Arachis hypogaea"),
    ("osa", "Oryza sativa"),
    ("mtr", "Medicago truncatula"),
]


# ── Load Arahy -> LOC crosswalk  (needed for ahf module enrichment) ────────
arahy_to_loc: dict = {}
if CROSSWALK_XLSX.exists():
    print(f"\n  Loading Arahy->LOC crosswalk: {CROSSWALK_XLSX.name} ...")
    xw = pd.read_excel(CROSSWALK_XLSX)
    arahy_col = loc_col = None
    for c in xw.columns:
        cl = str(c).lower()
        if "arahy" in cl or "phytozome" in cl or "gnm" in cl or cl == "gene":
            if arahy_col is None: arahy_col = c
        if ("loc" in cl and cl != "loc") or "ncbi" in cl or \
           cl in ("gene_id", "geneid", "locus", "loc_id"):
            if loc_col is None: loc_col = c
    if arahy_col is None: arahy_col = xw.columns[0]
    if loc_col is None:   loc_col   = xw.columns[1] if len(xw.columns) > 1 else xw.columns[0]
    print(f"    Columns: '{arahy_col}' -> '{loc_col}'")
    for _, row in xw.iterrows():
        a = str(row[arahy_col]).strip(); l = str(row[loc_col]).strip()
        if a and l and a != "nan" and l != "nan":
            arahy_to_loc[a] = l
    print(f"    {len(arahy_to_loc):,} Arahy->LOC mappings loaded")
else:
    print(f"\n  WARNING: crosswalk not found: {CROSSWALK_XLSX}")


# ── Helpers ────────────────────────────────────────────────────────────────
def _to_kegg_id(gene: str, org: str) -> str:
    """Convert annotation ID -> KEGG org: prefixed ID."""
    g = str(gene).strip()
    # Arachis: Arahy.XXXXXX -> LOCxxxxxxxxx -> strip LOC -> ncbi gene id
    if org == "ahf":
        g = arahy_to_loc.get(g, g)     # Arahy -> LOC ID
        if g.upper().startswith("LOC") and g[3:].split(".")[0].isdigit():
            g = g[3:]                   # LOCxxxxxxxxx -> xxxxxxxxx
        return f"ahf:{g}"
    # Rice: Os01g0100100 -> passed as-is; KEGG osa uses numeric ids
    #       (no direct mapping; we forward for completeness)
    if org == "osa":
        return f"osa:{g}"
    # Medicago: Medtr1g004950 -> passed as-is; KEGG mtr uses numeric ids
    if org == "mtr":
        return f"mtr:{g}"
    return f"{org}:{g}"


def _build_modules(communities: dict, org: str) -> dict:
    """Group genes by module id; convert to KEGG namespace."""
    modules: dict = {}
    n_conv = n_raw = 0
    for gene, mod_id in communities.items():
        kegg_id = _to_kegg_id(str(gene), org)
        modules.setdefault(mod_id, []).append(kegg_id)
        if org == "ahf" and str(gene) in arahy_to_loc:
            n_conv += 1
        else:
            n_raw += 1
    if org == "ahf":
        print(f"    ID mapping: {n_conv:,} Arahy->LOC converted, {n_raw:,} unchanged")
    return modules


def _simulate_modules_kegg(n_modules: int, genes_per_module: int,
                            bg_list: list, rng, org: str) -> dict:
    """
    Build simulated modules from KEGG background genes.
    Used for osa/mtr where annotation IDs don't map to KEGG numeric IDs.
    Sizes match the real module distribution for representative timing.
    """
    modules = {}
    for i in range(n_modules):
        n = min(genes_per_module, len(bg_list))
        modules[i] = list(rng.choice(bg_list, size=n, replace=False))
    return modules


def _run_enrichment_modules(modules_by_id: dict, gene_sets: dict,
                             background: set, source: str) -> tuple:
    """
    Time hypergeometric_enrichment across all modules.
    Returns (mean_t, sd_t, 0.0, n_enriched_modules, n_sig_total).
    """
    eligible = {k: v for k, v in modules_by_id.items() if len(v) >= 3}
    n_eligible = len(eligible)
    print(f"    {source}: {n_eligible} modules (>=3 genes)  |  "
          f"bg {len(background):,}  |  {len(gene_sets)} pathways")
    if n_eligible == 0:
        return 0.0, 0.0, 0.0, 0, 0

    all_times = []; total_sig = 0
    for genes in eligible.values():
        t0 = time.perf_counter()
        res = hypergeometric_enrichment(genes, gene_sets, background)
        all_times.append(time.perf_counter() - t0)
        if hasattr(res, "table") and not res.table.empty:
            total_sig += int((res.table["q_value"] < 0.05).sum())

    mean_t = float(np.mean(all_times))
    sd_t   = float(np.std(all_times, ddof=1)) if len(all_times) > 1 else 0.0
    print(f"    + {source}: {n_eligible} modules  |  mean {mean_t:.4f} s/module  |  "
          f"total {sum(all_times):.2f} s  |  {total_sig} sig. terms (q<0.05)")
    return mean_t, sd_t, 0.0, n_eligible, total_sig


# ── Storage for per-organism KEGG data ────────────────────────────────────
kegg_data = {}   # org -> {"gene_sets": dict, "bg": set, "n_pathways": int}
_rng_kegg = np.random.default_rng(42)


# ════════════════════════════════════════════════════════════════════════════
# 5a. First-time pathway fetch  (all 3 organisms, clear cache first)
# ════════════════════════════════════════════════════════════════════════════
print("\n  5a. First-time KEGG pathway fetch  (network, all 3 organisms)")

for org, org_label in KEGG_ORGS:
    for suffix in (f"{org}_paths.tsv", f"{org}_links.tsv"):
        p = _kegg_cache / suffix
        if p.exists():
            p.unlink()
            print(f"    Cleared cache: {suffix}")

    def _fetch_fresh(o=org):
        return _kegg_client.list_pathways(o), _kegg_client.pathway_genes(o)

    print(f"\n    Organism: {org} ({org_label})")
    mean_t, sd_t, peak_mb, (paths, raw_genes) = run_timed(
        f"KEGG fetch (network) [{org}]", _fetch_fresh, n_reps=1,
    )
    n_paths  = len(paths)
    n_assoc  = sum(len(v) for v in raw_genes.values())
    print(f"    -> {n_paths} pathways, {n_assoc:,} gene-pathway associations")

    # Normalise: strip 'path:' prefix and build background set
    gene_sets = {k.split("path:", 1)[-1]: v for k, v in raw_genes.items()}
    bg = set()
    for gs in gene_sets.values():
        bg.update(gs)
    kegg_data[org] = {"gene_sets": gene_sets, "bg": bg, "n_pathways": n_paths}

    record(
        function  = "KEGG Pathway Fetch (network)",
        dataset   = f"{org_label} ({org})",
        parameter = "list/pathway + link/pathway; REST API",
        genome_mb = "", n_input="", n_output=n_paths,
        mean_t=mean_t, sd_t=sd_t, peak_mb=peak_mb,
    )


# ════════════════════════════════════════════════════════════════════════════
# 5b. Cached pathway fetch  (all 3 organisms, 5 reps)
# ════════════════════════════════════════════════════════════════════════════
print("\n  5b. Cached KEGG pathway fetch  (disk, 5 reps, all 3 organisms)")

for org, org_label in KEGG_ORGS:
    kd = kegg_data[org]

    def _fetch_cached(o=org):
        return _kegg_client.list_pathways(o), _kegg_client.pathway_genes(o)

    print(f"\n    Organism: {org}")
    mean_t_c, sd_t_c, peak_mb_c, _ = run_timed(
        f"KEGG fetch (cached) [{org}]", _fetch_cached, n_reps=5,
    )
    print(f"    -> {mean_t_c:.4f} s (cached)  |  {kd['n_pathways']} pathways")
    record(
        function  = "KEGG Pathway Fetch (cached)",
        dataset   = f"{org_label} ({org})",
        parameter = "list/pathway + link/pathway; disk cache",
        genome_mb = "", n_input="", n_output=kd["n_pathways"],
        mean_t=mean_t_c, sd_t=sd_t_c, peak_mb=peak_mb_c,
    )


# ════════════════════════════════════════════════════════════════════════════
# 5c. ID Convert  (ahf only — Arachis NCBI Gene IDs -> ahf: namespace)
# ════════════════════════════════════════════════════════════════════════════
print("\n  5c. ID Convert benchmark  (KEGGClient.convert_to_kegg, organism: ahf)")

_hit_locs: list = []
if HITS_CSV.exists():
    _hits_df = pd.read_csv(HITS_CSV)
    _id_col  = next((c for c in _hits_df.columns
                     if "record" in c.lower() or "gene" in c.lower()),
                    _hits_df.columns[0])
    for _g in _hits_df[_id_col].dropna().unique():
        _s = str(_g).strip()
        if _s.upper().startswith("LOC") and _s[3:].split(".")[0].isdigit():
            _hit_locs.append(_s[3:])
        elif _s.isdigit():
            _hit_locs.append(_s)
    print(f"    {len(_hit_locs):,} numeric gene IDs from {HITS_CSV.name}")
else:
    print(f"    HITS_CSV not found; using 200 synthetic IDs")
    _hit_locs = [str(110000000 + i) for i in range(200)]

for _n_ids in [50, min(200, len(_hit_locs))]:
    _sample = _hit_locs[:_n_ids]
    print(f"\n    ID Convert: {_n_ids} NCBI Gene IDs -> ahf: namespace  (1 rep)")
    mean_t_id, sd_t_id, peak_mb_id, _id_map = run_timed(
        f"ID Convert n={_n_ids}",
        _kegg_client.convert_to_kegg, "ahf", _sample,
        n_reps=1,
    )
    print(f"    -> {len(_id_map)}/{_n_ids} IDs mapped to ahf: namespace")
    record(
        function  = "KEGG ID Convert",
        dataset   = "Arachis hypogaea (ahf)",
        parameter = f"{_n_ids} NCBI Gene IDs -> ahf: namespace; /conv REST",
        genome_mb = "", n_input=_n_ids, n_output=len(_id_map),
        mean_t=mean_t_id, sd_t=sd_t_id, peak_mb=peak_mb_id,
        throughput=round(_n_ids / mean_t_id, 1) if mean_t_id > 0 else "",
        unit="IDs/s",
    )


# ════════════════════════════════════════════════════════════════════════════
# 5d. KEGG enrichment per Louvain module  (all 3 organisms)
#   - ahf : real Arahy->LOC converted gene IDs
#   - osa : KEGG-native gene IDs sampled from osa background (performance proxy)
#   - mtr : KEGG-native gene IDs sampled from mtr background (performance proxy)
# ════════════════════════════════════════════════════════════════════════════
print("\n  5d. KEGG enrichment per Louvain module  (all 3 organisms)")

for sp in COEXPR_SPECIES:
    key  = sp["key"]
    org  = sp["kegg_org"]
    comm = coexpr_communities.get(key)
    if not comm:
        print(f"\n  SKIP 5d [{sp['label']}] -- Louvain communities unavailable")
        continue

    kd         = kegg_data[org]
    gene_sets  = kd["gene_sets"]
    background = kd["bg"]

    print(f"\n    {sp['label']}  (organism: {org})")

    if org == "ahf":
        # Real Arahy -> LOC -> ahf: IDs
        modules = _build_modules(comm, org)
        source  = "Louvain (real Arahy->LOC IDs)"
    else:
        # Simulate module gene lists from KEGG background (representative timing)
        real_modules = _build_modules(comm, org)
        n_mod = sum(1 for v in real_modules.values() if len(v) >= 3)
        avg_sz = max(3, int(np.mean([len(v) for v in real_modules.values()])))
        bg_list = list(background)
        modules = _simulate_modules_kegg(n_mod, avg_sz, bg_list, _rng_kegg, org)
        source  = f"Louvain (KEGG-native IDs; n_mod={n_mod}, avg_size={avg_sz})"
        print(f"    NOTE: {org} uses KEGG-native gene IDs (numeric) for timing only")

    mean_t, sd_t, peak_mb, n_mod, n_sig = _run_enrichment_modules(
        modules, gene_sets, background, source,
    )
    record(
        function  = "KEGG Enrichment (Louvain modules)",
        dataset   = f"{sp['label']} ({org})",
        parameter = f"{n_mod} modules, {len(gene_sets)} pathways",
        genome_mb = "", n_input=n_mod, n_output=n_sig,
        mean_t=mean_t, sd_t=sd_t, peak_mb=peak_mb,
        throughput=round(1 / mean_t, 2) if mean_t > 0 else "",
        unit="modules/s",
    )


# ════════════════════════════════════════════════════════════════════════════
# 5e. KEGG enrichment per k-means cluster  (all 3 organisms)
# ════════════════════════════════════════════════════════════════════════════
print("\n  5e. KEGG enrichment per k-means cluster  (all 3 organisms)")

for sp in COEXPR_SPECIES:
    key  = sp["key"]
    org  = sp["kegg_org"]
    comm = kmeans_communities.get(key)
    if not comm:
        print(f"\n  SKIP 5e [{sp['label']}] -- k-means communities unavailable")
        continue

    kd         = kegg_data[org]
    gene_sets  = kd["gene_sets"]
    background = kd["bg"]

    print(f"\n    {sp['label']}  (organism: {org})")

    if org == "ahf":
        modules = _build_modules(comm, org)
        source  = "k-means (real Arahy->LOC IDs)"
    else:
        real_modules = _build_modules(comm, org)
        n_mod  = sum(1 for v in real_modules.values() if len(v) >= 3)
        avg_sz = max(3, int(np.mean([len(v) for v in real_modules.values()])))
        bg_list = list(kd["bg"])
        modules = _simulate_modules_kegg(n_mod, avg_sz, bg_list, _rng_kegg, org)
        source  = f"k-means (KEGG-native IDs; n_mod={n_mod}, avg_size={avg_sz})"
        print(f"    NOTE: {org} uses KEGG-native gene IDs (numeric) for timing only")

    mean_t, sd_t, peak_mb, n_mod, n_sig = _run_enrichment_modules(
        modules, gene_sets, background, source,
    )
    record(
        function  = "KEGG Enrichment (k-means clusters)",
        dataset   = f"{sp['label']} ({org})",
        parameter = f"{n_mod} clusters, {len(gene_sets)} pathways",
        genome_mb = "", n_input=n_mod, n_output=n_sig,
        mean_t=mean_t, sd_t=sd_t, peak_mb=peak_mb,
        throughput=round(1 / mean_t, 2) if mean_t > 0 else "",
        unit="modules/s",
    )


# ════════════════════════════════════════════════════════════════════════════
# 5f. Scaling test -- query size vs enrichment time  (all 3 organisms)
# ════════════════════════════════════════════════════════════════════════════
print("\n  5f. Scaling test: query size vs enrichment time  (all 3 organisms)")

for org, org_label in KEGG_ORGS:
    kd         = kegg_data[org]
    gene_sets  = kd["gene_sets"]
    background = kd["bg"]
    bg_list    = list(background)
    print(f"\n    Organism: {org}  |  bg {len(background):,}  |  {len(gene_sets)} pathways")

    for _n_q in [50, 200, 500, 2000]:
        _n_samp = min(_n_q, len(bg_list))
        _query  = list(_rng_kegg.choice(bg_list, size=_n_samp, replace=False))
        mean_t, sd_t, peak_mb, _ = run_timed(
            f"KEGG scaling {org} n={_n_q}", hypergeometric_enrichment,
            _query, gene_sets, background, n_reps=5,
        )
        record(
            function  = "KEGG Enrichment (scaling)",
            dataset   = f"{org_label} ({org})",
            parameter = f"{_n_samp} query genes, {len(gene_sets)} pathways, "
                        f"{len(background):,} background",
            genome_mb = "", n_input=_n_samp, n_output="",
            mean_t=mean_t, sd_t=sd_t, peak_mb=peak_mb,
            throughput=round(_n_samp / mean_t, 1) if mean_t > 0 else "",
            unit="genes/s",
        )


# ══════════════════════════════════════════════════════════════════════════
# SAVE RESULTS
# ══════════════════════════════════════════════════════════════════════════
print("\n" + "-" * 70)
print("  Saving results ...")

df = pd.DataFrame(results)
for col, val in [
    ("python_version", PY_VER),
    ("ram_gb",         RAM_GB),
    ("cpu_threads",    CPU_THREADS),
    ("cpu_cores",      CPU_CORES),
    ("cpu",            CPU_NAME),
    ("os",             OS_NAME),
    ("timestamp",      datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
]:
    df.insert(0, col, val)

df.to_csv("benchmark_results.csv", index=False)
print("  Saved -> benchmark_results.csv")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    # Color scheme: one colour per function family
    FILLS = {
        "Promoter Extraction":                   "1A5E38",  # dark green
        "Motif Search":                          "1F4E79",  # dark blue
        "Expression Normalisation":              "6C3483",  # purple
        "Correlation Matrix":                    "6C3483",
        "Co-expression Clustering (Louvain)":    "6C3483",
        "k-means Elbow Method":                  "784212",  # brown
        "k-means Clustering":                    "784212",
        "KEGG Pathway Fetch (network)":          "7B3F00",  # dark brown
        "KEGG Pathway Fetch (cached)":           "A04000",  # medium brown
        "KEGG ID Convert":                       "B7460A",  # orange-brown
        "KEGG Enrichment (Louvain modules)":     "7B3F00",
        "KEGG Enrichment (k-means clusters)":    "7B3F00",
        "KEGG Enrichment (scaling)":             "7B3F00",
    }

    wb = Workbook()

    # System Info sheet
    ws0 = wb.active; ws0.title = "System Info"
    sysinfo = [
        ("Benchmark Date",  datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("OS",              OS_NAME),
        ("CPU",             CPU_NAME),
        ("Physical Cores",  CPU_CORES),
        ("Logical Threads", CPU_THREADS),
        ("RAM",             f"{RAM_GB} GB"),
        ("Python",          PY_VER),
        ("Replicates",      "3-5 per condition"),
        ("Time metric",     "time.perf_counter() wall time"),
        ("RAM metric",      "psutil RSS delta (no timing overhead)"),
        ("Promoter length", f"{PROMOTER_LEN} bp"),
        ("Species (coexpr)","Arachis gnm1, Oryza sativa IRGSP-1.0, Medicago Mt5.0"),
        ("KEGG organisms",  "ahf (Arachis), osa (Oryza), mtr (Medicago)"),
    ]
    for ri, (k, v) in enumerate(sysinfo, 1):
        ws0.cell(ri, 1, k).font = Font(bold=True, name="Arial", size=10)
        ws0.cell(ri, 2, v).font = Font(name="Arial", size=10)
    ws0.column_dimensions["A"].width = 22
    ws0.column_dimensions["B"].width = 60

    # Results sheet
    ws1 = wb.create_sheet("Results")
    HDR = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    DAT = Font(name="Arial", size=9)

    meta_cols = {"timestamp","os","cpu","cpu_cores","cpu_threads","ram_gb","python_version"}
    cols = [c for c in df.columns if c not in meta_cols]
    for ci, c in enumerate(cols, 1):
        cell = ws1.cell(1, ci, c)
        cell.font = HDR
        cell.fill = PatternFill("solid", start_color="1F4E79")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws1.row_dimensions[1].height = 30

    sub_df = df[cols]
    for ri, row in enumerate(sub_df.values.tolist(), 2):
        fn = row[0]
        fc = FILLS.get(fn, "F0F0F0")
        tc = "FFFFFF" if fc != "F0F0F0" else "1A1A1A"
        for ci, val in enumerate(row, 1):
            cell = ws1.cell(ri, ci, val)
            cell.font = Font(name="Arial", size=9, color=tc)
            cell.fill = PatternFill("solid", start_color=fc)
            cell.alignment = Alignment(
                horizontal="right" if ci > 1 else "left"
            )

    for ci, c in enumerate(cols, 1):
        ws1.column_dimensions[get_column_letter(ci)].width = max(14, len(c) + 2)
    ws1.freeze_panes = "A2"

    wb.save("benchmark_results.xlsx")
    print("  Saved -> benchmark_results.xlsx")
except Exception as e:
    print(f"  Excel export skipped: {e}")

print("\n" + "=" * 70)
print("  BENCHMARK COMPLETE")
print(f"  Total functions : {df['function'].nunique()}")
print(f"  Total rows      : {len(df)}")
print("=" * 70)
